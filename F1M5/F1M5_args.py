import argparse
import openpyxl
import shutil
import datetime


now = datetime.datetime.now()
timestamp_today = now.strftime("%m%d")
yesterday = datetime.date.today() - datetime.timedelta(days=1)
timestamp_yesterday = yesterday.strftime("%m%d")

parser = argparse.ArgumentParser(description='Process filenames and shift.')
parser.add_argument('shift', choices=['M', 'S'], help='Shift (M or S)')
parser.add_argument('game_test', type=str, help='Game test filename')
parser.add_argument('web_test', type=str, help='Web test filename')
args = parser.parse_args()

shift = args.shift
game_test = args.game_test
web_test = args.web_test

wb_game = openpyxl.load_workbook(game_test)
sheet_game = wb_game.active

wb_web = openpyxl.load_workbook(web_test)
sheet_web = wb_web.active

if shift == 'M':
    filename = 'F1M5_' + timestamp_yesterday + '_2100-' + timestamp_today + '_0900'
elif shift == 'S':
    filename = 'F1M5_' + timestamp_today + '_0900-' + timestamp_today + '_2100'
source_file = 'F1M5.xlsx'
des_file = filename + ',xlsx'
shutil.copy2('F1M5.xlsx', filename + '.xlsx')

wb2 = openpyxl.load_workbook(filename + '.xlsx')
sheet2 = wb2['F1M5_Game_Test 0900-2100']
sheet3 = wb2['F1M5 0900-2100']

city_isp_map = {
    'Hyderabad_TATA': 5,
    'Hyderabad_Airtel Ltd.': 6,
    'Bangalore_TATA': 7,
    'Bangalore_Airtel Ltd.': 8,   
    'Bangalore_Vodafone': 9,      
    'Delhi_Reliance': 10,
    'Delhi_JIO': 11,
    'Delhi_TATA': 12,    
}

city_isp_map_web = {
    'Hyderabad_TATA': 4,
    'Hyderabad_Airtel Ltd.': 5,
    'Bangalore_TATA': 6,
    'Bangalore_Airtel Ltd.': 7,   
    'Bangalore_Vodafone': 8,      
    'Delhi_Reliance': 9,
    'Delhi_JIO': 10,
    'Delhi_TATA': 11,    
}

step_map = {
    'Step1-Homepage': 4,
    'Step2-Login': 5,
    'Step3 Saba': 10,
    'Step3-Desktop-Betb2b': 13,
    'Step4-BTI': 6,
    'Step5-Cricket': 7,
    'Step6-Evolution': 8,
    'Step7 mobile-Betb2b': 9,
}

test_map = {
    'Mobile - fun1005.com': 7,
    'Mobile - fun88in.co': 6,
    'Mobile - fun88inr.com': 8,
    'Mobile - fun9262.com': 5,
    'Mobile - fun9915.com': 4   
}

for i in range(13, 78):
    try:
        value1_City_ISP = sheet_game.cell(row=i, column=2).value + '_' + sheet_game.cell(row=i, column=3).value
        value1_StepName = sheet_game.cell(row=i, column=4).value
    
        if value1_City_ISP in city_isp_map and value1_StepName in step_map:
            sheet2.cell(row=step_map[value1_StepName], column=city_isp_map[value1_City_ISP]).value = sheet_game.cell(row=i, column=6).value
        else:
            print("game_test some data missing")
    except:
        pass

for c in range(16, 56):
    try:
        value2_City_ISP = sheet_web.cell(row=c, column=2).value + '_' + sheet_web.cell(row=c, column=3).value
        value2_StepName = sheet_web.cell(row=c, column=4).value
        if value2_City_ISP in city_isp_map_web and value2_StepName in test_map:
            sheet3.cell(row=test_map[value2_StepName], column=city_isp_map_web[value2_City_ISP]).value = sheet_web.cell(row=c, column=7).value
        else:
            print("web_test some data missing")
    except:
        pass

# 保存第二個Excel文件
wb2.save(filename + '.xlsx')
print("done")
input("Press enter to exit...")